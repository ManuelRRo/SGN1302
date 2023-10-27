
from typing import Any
from django.db.models.query import QuerySet
from django.shortcuts import get_object_or_404, render, redirect
from .models import Evaluacion, Evaluacionalumno, Alumno, Gradoseccion, Docente, Materia, Gradoseccionmateria
from django.shortcuts import render, redirect
from .models import Evaluacion, Evaluacionalumno, Alumno, Gradoseccion, Docente, Materia, Gradoseccionmateria, Trimestre, Promediomateria
from .forms import EvaluacionForm, EvaluacionAlumnoForm, DocenteForm, AlumnoForm, TrimestreActualizarForm, EvaluacionEditarForm, TrimestreForm
from aplicaciones.usuarios.forms import RegisterUserForm
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View, TemplateView
from django.urls import reverse_lazy, reverse
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.forms import UserCreationForm,UserChangeForm,SetPasswordForm
from openpyxl import Workbook
from django.http.response import HttpResponse
from aplicaciones.sistemadenotas.filters import EvaluacionFilter
from datetime import datetime
from openpyxl import load_workbook
import matplotlib
matplotlib.use('Agg')
from matplotlib import pyplot as plt
from io import BytesIO
import base64
from asgiref.sync import sync_to_async
from django.db.models import Count, Q 
from operator import itemgetter
from decimal import Decimal

def asignacionClases(request):
    context = {}
    existenciaDocente = Docente.objects.filter(numidentificacion=request.user.username).exists()
    # Al no ser admin se cumple la HU-01
    if existenciaDocente:
        docente = Docente.objects.get(numidentificacion=request.user.username)
        materia = Materia.objects.filter(id_docente=docente)
        grado_seccion_materia = Gradoseccionmateria.objects.filter(id_materia__in=materia)
        gradoseccion = Gradoseccion.objects.filter(gradoseccionmateria__id_materia__id_docente=docente).distinct()
        context['grado_seccion'] = gradoseccion
        context["grado_seccion_materia"] = grado_seccion_materia
    return context

# HU-01 Listar Grados asignados | Materias impartidas
# Posee dos comportamientos:
#   - Rol profesor -> cumple HU-01
#   - Rol Administrador -> no se ejecuta HU-01
@login_required()
def home(request):
    context = {}
    context = asignacionClases(request)
    
    return render (request,'home/inicio.html',context)


# HU-02 Listar Evaluaciones de Grado
# De acuerdo a la materia seleccionada de ese grado
class ListarEvaluacionesGrados(ListView):
    model = Evaluacion
    # context_object_name = 'evas'
    template_name = 'evaluacion/evaluacion.html'

    def get_queryset(self):
        self.idgrado = self.kwargs["idgrado"]
        if self.idgrado != None:
            return self.model.objects.filter(id_gradoseccionmateria=self.idgrado)
        return self.model.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        evaluacion_filter = EvaluacionFilter(self.request.GET,queryset = self.get_queryset())
        # Semejante al home: para el navbar
        context = asignacionClases(self.request)
        # ------------------------------------
        context["filter_form"] = evaluacion_filter.form
        context["evas"] = evaluacion_filter.qs
 
        return context


# # VISTAS HU-03 y HU-09
# def CrearEvaluacionAlumno(request):
#     submitted = False
#     if request.method == "POST":
#         # form = EvaluacionForm(request.POST)
#         form = EvaluacionForm(request.POST)
#         if form.is_valid():
#             evaluacion = form.save()  # contiene los datos de la evaluacion que se acaba de crear
#             # grado = form.cleaned_data['id_gradoseccionmateria'].id_gradoseccion.id_gradoseccion
#             if evaluacion is not None:
#                 grado = evaluacion.id_gradoseccionmateria.id_gradoseccion.id_gradoseccion
#                 alumno = Alumno.objects.filter(id_gradoseccion=grado)
#                 for e in alumno:
#                     evaalumno = Evaluacionalumno.objects.create(
#                         id_evaluacion=evaluacion, id_alumno=e, nota=0.0)
#                 return HttpResponseRedirect('/estudiante/crear-eva-est?submitted=True')
#     else:
#         form = EvaluacionForm
#         # USER SUBMITTER THE FORM
#         if 'submitted' in request.GET:
#             submitted = True
#     context = {
#         'form': form,
#         'submitted': submitted,
#     }
#     return render(request, 'estudiante/crear-evaluacion.html', context)


# HU-04 - HU-05

# Permite listar los alumnos de esta evalucion
# a la vez actualizar y registrar
class ListarEvaluacionesAlumnos(View):
    model = Evaluacionalumno
    form_class = EvaluacionAlumnoForm
    template_name = 'estudiante/listar-evas-alumnos.html'

    def get_queryset(self):
        self.evaluacion = self.kwargs["idEvaluacion"]
        self.alumnos = Evaluacionalumno.objects.filter(
            id_evaluacion=self.evaluacion)
        return self.alumnos

    def get_context_data(self, **kwargs):
        contexto = {}
        contexto = asignacionClases(self.request)
        contexto['estudiantes'] = self.get_queryset()
        contexto['form'] = self.form_class
        return contexto

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, self.get_context_data())

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            try:
                idevaluacion = request.POST['idEvaluacion']
                print(idevaluacion)
                self.Evaluacion = Evaluacion.objects.get(
                    id_evaluacion=idevaluacion)
                idAlumno = request.POST.get('idAlumno', None)
                nota = request.POST['nota']
                if idAlumno is not None and idAlumno != '':
                    self.Alumno = Alumno.objects.get(id_alumno=idAlumno)
                    self.EvaluacionAlumno = Evaluacionalumno.objects.get(
                        id_evaluacion=self.Evaluacion, id_alumno=self.Alumno)
                    self.EvaluacionAlumno.nota = nota
                    self.EvaluacionAlumno.save()
                    messages.success(request, 'Calificado correctamente')
            except Exception:
                messages.error(request, 'Ocurrio un error')

        # ID de evaluación que deseas pasar a la URL
        return redirect('sgn_app:list_evas_not', idEvaluacion=idevaluacion)


# HU-07: Mostrar Notas de Alumnos de todas las evaluaciones
# Mostrarse baso en el archivo de excel que proporcionaron en el centro escolar.
def ver_Promedios(request, idgrado, idtrimestre):

    nulos = False
    
    gradoseccion = Gradoseccionmateria.objects.get(id_gradoseccionmateria=idgrado)
    
    alumnos = Alumno.objects.filter(id_gradoseccion=gradoseccion.id_gradoseccion,estado='1')
    
    evaluacionalumno = Evaluacionalumno.objects.filter(id_alumno__id_gradoseccion=gradoseccion.id_gradoseccion, 
                                                        id_evaluacion__id_trimestre=idtrimestre, 
                                                        id_evaluacion__id_gradoseccionmateria=gradoseccion.id_gradoseccionmateria,
                                                        id_evaluacion__estado=1).order_by('id_evaluacion')
    
    evaluacion_ids = evaluacionalumno.values_list('id_evaluacion', flat=True)
    
    evaluaciones = Evaluacion.objects.filter(id_evaluacion__in=evaluacion_ids).filter(id_trimestre=idtrimestre)
    
    materia = Materia.objects.get(id_materia=gradoseccion.id_materia.id_materia)
    
    trimestre = Trimestre.objects.get(id_trimestre=idtrimestre)

    promedios = []
    aprobados = 0
    reprobados = 0
    for alumno in alumnos:
        notas = []
        for evaluacion in evaluaciones:
            evaluacion_alumno = evaluacionalumno.filter(
                id_alumno=alumno, id_evaluacion=evaluacion).first()
            if evaluacion_alumno:
                try:
                    nota_ponderada = evaluacion_alumno.nota * \
                    (evaluacion.porcentaje / 100)
                    
                except:
                    nota_ponderada = 0
                    nulos = True
                notas.append(nota_ponderada)
        promedio = round(sum(notas), 2)
        if(promedio >= 5.0):
            aprobados += 1
        else:
            reprobados += 1
        promedios.append({'alumno': alumno, 'promedio': promedio})


    contexto = {
        'gradoseccion': gradoseccion,
        'evaluaciones': evaluaciones,
        'evaluacionesalumno': evaluacionalumno,
        'materia': materia,
        'trimestre': trimestre,
        'alumnos': alumnos,
        'promedios': promedios,
        'aprobados': aprobados,
        'reprobados': reprobados,
        'nulos':nulos,
    }

    contexto.update(asignacionClases(request))
    

    return render(request, 'calificaciones/verPromedios.html', contexto)


# Codigo HU-08 Generar archivo de excel de notas trimestrales 
class ReporteDeNotasExcel(TemplateView):
    def post(self, request, *args, **kwargs):
        # Obtiene los valores enviados desde el formulario
        comentarios = {}
        valores_promedio = {}
        grado = request.POST.get('grado', '')
        trimestre = request.POST.get('trimestre', '')

        for key, value in request.POST.items():
            if key.startswith('opcional_'):
                alumno_id = key.replace('opcional_', '')
                comentarios[alumno_id] = value
            elif key.startswith('promedio_'):
                alumno_id = key.replace('promedio_', '')
                valores_promedio[alumno_id] = value

        for alumno_id, valor in comentarios.items():
            # Accede a los valores por el ID del alumno
            alumno = Alumno.objects.get(id_alumno=alumno_id)

        for alumno_id, valor in valores_promedio.items():
            # Accede a los valores por el ID del alumno
            alumno = Alumno.objects.get(id_alumno=alumno_id)

        # Código para generar y devolver el archivo Excel
        return self.get(request, comentarios=comentarios, trimestre=trimestre, grado=grado, valores_promedio=valores_promedio, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtén los valores opcionales de los argumentos de la vista
        comentarios = kwargs.get('comentarios')
        notas_finales = kwargs.get('valores_promedio')
        trimestre = kwargs.get('trimestre')
        grado = kwargs.get('grado')

        alumnos = Alumno.objects.none()  # Inicializa una consulta vacía de Alumno

        for alumno_id, valor in comentarios.items():
            # Filtra los alumnos según los valores opcionales
            alumnos |= Alumno.objects.filter(id_alumno=alumno_id)
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'nie'
        ws['B1'] = 'calificacion'
        ws['C1'] = 'fecha'
        ws['D1'] = 'observacion'

        cont = 2
        numero = 1

        for alumno in alumnos:
            ws.cell(row=cont, column=1).value = int(alumno.nie)
            alumno_id = str(alumno.id_alumno)
            if alumno_id in notas_finales:
                nota_final = notas_finales[alumno_id]
                ws.cell(row=cont, column=2).value = float(nota_final)
            ws.cell(row=cont, column=3).value = datetime.now().strftime(
                '%d/%m/%Y')
            alumno_id = str(alumno.id_alumno)
            if alumno_id in comentarios:
                comentario = comentarios[alumno_id]
                ws.cell(row=cont, column=4).value = comentario
            else:
                ws.cell(row=cont, column=4).value = ""
            cont += 1
            numero += 1

        nombre_archivo = "Notas " + grado + " " + trimestre+".xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        content = "attachment; filename={0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response

# HU10 Cambiar Rol de Usuario


class cambiarRolListView(ListView):
    model = Docente
    template_name = 'docente/cambiar_Rol.html'
    context_object_name = 'docentes'
    queryset = model.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['users'] = User.objects.all()
        context.update(asignacionClases(self.request))
        return context


def Cambiar_Rol(request, nombreUsuario):
    try:
        # Recupera el registro por el nombre de usuario
        registro = User.objects.get(username=nombreUsuario)

        # Cambia el estado is_superuser
        registro.is_superuser = not registro.is_superuser
        registro.save()
        messages.success(request, 'Rol cambiado con éxito')
    except User.DoesNotExist:
        pass
    return redirect('/cambiar_Rol/')
# HU-14: Reporte de Alumnos Aprobados/Reprobados
def graficoAR(request, aprobados, reprobados, gradoseccionmateria):
    contexto = {}
    contexto.update(asignacionClases(request))
    contexto['aprobados'] = aprobados
    contexto['reprobados'] = reprobados
    contexto['gradoseccionmateria'] = Gradoseccionmateria.objects.get(id_gradoseccionmateria = gradoseccionmateria)
    return render(request, 'administracion/graficoAR.html', contexto)

#HU-15: Reporte de Alumnos Masculinos/Femeninos
def generar_grafico_barra(data, title):
    grados = list(data.keys())
    alumnos_masculinos = [item['M'] for item in data.values()]
    alumnos_femeninos = [item['F'] for item in data.values()]

    x = range(len(grados))
    width = 0.4

    plt.bar(x, alumnos_masculinos, width=width, label='Masculino')
    plt.bar([pos + width for pos in x], alumnos_femeninos, width=width, label='Femenino')

    plt.xlabel("Grados")
    plt.ylabel("Número de alumnos")
    plt.title(title)
    plt.xticks([pos + width / 2 for pos in x], grados)
    plt.legend()

    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.read()).decode('utf-8')

    plt.close()
    return image_base64

def generar_graficos_por_ciclo(ciclo):
    grados_secciones = Gradoseccion.objects.filter(id_grado__in=range(ciclo[0], ciclo[1]), id_seccion__in=range(1, 5))
    data = []

    for gradoseccion in grados_secciones:
        gradoseccion_text = f"{gradoseccion.id_grado.grado} {gradoseccion.id_seccion.seccion}"
        alumnos = Alumno.objects.filter(id_gradoseccion=gradoseccion)
        total_masculinos = alumnos.filter(sexo='M').count()
        total_femeninos = alumnos.filter(sexo='F').count()
        data.append({
            'gradoseccion': gradoseccion_text,
            'masculinos': total_masculinos,
            'femeninos': total_femeninos
        })
    data.sort(key=itemgetter('gradoseccion'))
    ciclo_nombre = ""
    if ciclo == (1, 4):
        ciclo_nombre = "Primer Ciclo"
    elif ciclo == (4, 7):
        ciclo_nombre = "Segundo Ciclo"
    elif ciclo == (7, 10):
        ciclo_nombre = "Tercer Ciclo"
        
        def custom_sort_key(item):
            grado, seccion = item['gradoseccion'].rsplit(' ', 1)
            orden_grados = {"Septimo Grado": 1, "Octavo Grado": 2, "Noveno Grado": 3}
            return (orden_grados.get(grado, 0), seccion)

        data.sort(key=custom_sort_key)

    grados_secciones = [item['gradoseccion'] for item in data]
    masculinos = [item['masculinos'] for item in data]
    femeninos = [item['femeninos'] for item in data]

    ancho_barras = 0.2
    posicion_barras_femeninos = [pos + ancho_barras for pos in range(len(grados_secciones))]

    fig = plt.figure(figsize=(12, 6))
    plt.subplots_adjust(left=0.1, right=0.9, top=0.9, bottom=0.1)
    bars1 = plt.bar(range(len(grados_secciones)), masculinos, width=ancho_barras, label="Masculinos")
    bars2 = plt.bar(posicion_barras_femeninos, femeninos, width=ancho_barras, label="Femeninos")
    plt.grid(True, axis='y', linestyle='--', alpha=0.7)

    plt.xticks(range(len(grados_secciones)), grados_secciones, rotation=0)
    plt.title(f"Población Estudiantil de {ciclo_nombre}")

    plt.xlabel("Grado y Sección")
    plt.ylabel("Número de alumnos")

    plt.legend(loc='upper right')

    buffer = BytesIO()
    plt.savefig(buffer, format='png')
    buffer.seek(0)
    image_base64 = base64.b64encode(buffer.read()).decode('utf-8')

    plt.close()
    return image_base64

@sync_to_async
def graficos(request):
    image_base64_1 = generar_graficos_por_ciclo((1, 4))
    image_base64_2 = generar_graficos_por_ciclo((4, 7))
    image_base64_3 = generar_graficos_por_ciclo((7, 10))

    context = {
        'image_base64_1': image_base64_1,
        'image_base64_2': image_base64_2,
        'image_base64_3': image_base64_3
    }
    return render(request, 'administracion/graficos.html', context)
# HU-21: Insertar Alumnos
class CrearAlumno(CreateView):
    form_class = AlumnoForm
    template_name = 'estudiante/crear-alumnos.html'
    success_url = reverse_lazy('sgn_app:home')


class HabDeshabiAlumno(ListView):
    model = Alumno
    template_name = 'estudiante/hab-desh.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        id = self.kwargs['id']
        context['seccion'] = Gradoseccion.objects.get(id_gradoseccion=id)
        return context

    def get_queryset(self):
        id = self.kwargs['id']
        alumnos = Alumno.objects.filter(
            id_gradoseccion=id
        )
        return alumnos


def habilitar(request, id, idAlumno):
    alumno = Alumno.objects.get(id_alumno=idAlumno)
    alumno.estado = "1"
    alumno.save()
    messages.success(request, 'Alumno habilitado correctamente')
    return redirect(f'/habilitarDeshabilitarAlumno/{id}/')


def deshabilitar(request, id, idAlumno):
    alumno = Alumno.objects.get(id_alumno=idAlumno)
    alumno.estado = "0"
    alumno.save()
    messages.success(request, 'Alumno deshabilitado correctamente')
    return redirect(f'/habilitarDeshabilitarAlumno/{id}/')




# HU-22: Editar Alumno
class ActualizarAlumno(UpdateView):
    model = Alumno
    form_class = AlumnoForm
    template_name = 'estudiante/listar_alumnoGradoSeccion.html'

    def get_context_data(self,**kwargs):
        context = super().get_context_data(**kwargs)
        self.gradoseccion = self.kwargs["id_gradoseccion"]
        grado_del_docente = Gradoseccion.objects.get(id_gradoseccion=self.kwargs["id_gradoseccion"])
        
        self.alumnos = Alumno.objects.filter(id_gradoseccion=self.gradoseccion)
        context['estudiantes'] = self.alumnos
        context['grado_docente'] = grado_del_docente
        context.update(asignacionClases(self.request))
        return context
    
    def get_success_url(self):
        grado = self.kwargs['id_gradoseccion']
        url = reverse('sgn_app:ListarAlumno', kwargs={'id_gradoseccion': grado})
        return url
    
# HU-24: Listar Alumnos

class ListarAlumno(View):
    model = Alumno
    form_class = AlumnoForm
    template_name = 'estudiante/listar_alumnoGradoSeccion.html'

    def get_queryset(self):
        self.gradoseccion = self.kwargs["id_gradoseccion"]
        self.alumnos = Alumno.objects.filter(id_gradoseccion=self.gradoseccion)
        return self.alumnos

    def get_context_data(self, **kwargs):
        contexto = {}
        # Semejante al home: para el navbar
        docente = Docente.objects.get(numidentificacion=self.request.user.username)
        materia = Materia.objects.filter(id_docente=docente)
        grado_seccion_materia = Gradoseccionmateria.objects.filter(id_materia__in=materia)
        gradoseccion = Gradoseccion.objects.filter(gradoseccionmateria__id_materia__id_docente=docente).distinct()
        # ------------------------------------
        #nueva query de grados
        grado_del_docente = Gradoseccion.objects.get(id_gradoseccion=self.kwargs["id_gradoseccion"])
        # ------------------------------------
        contexto['grado_seccion'] = gradoseccion
        contexto["grado_seccion_materia"] = grado_seccion_materia
        contexto['estudiantes'] = self.get_queryset()
        contexto['grado_docente'] = grado_del_docente
        contexto['form'] = AlumnoForm()
        return contexto

    def get(self, request, *args, **kwargs):
        request.session['datos_archivo'] = None
        return render(request, self.template_name, self.get_context_data())

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST)
        if form.is_valid():
            alumno = form.save()
            numEvaluaciones = Evaluacionalumno.objects.filter(id_alumno=alumno).count()
            if numEvaluaciones == 0:
                gradoseccionmaterias = Gradoseccionmateria.objects.filter(id_gradoseccion=alumno.id_gradoseccion)
                for gradoseccionmateria in gradoseccionmaterias:
                    evaluaciones = Evaluacion.objects.filter(id_gradoseccionmateria=gradoseccionmateria)
                    for evaluacion in evaluaciones:
                        evaluacion_alumno = Evaluacionalumno.objects.create(id_evaluacion=evaluacion,
                                                                            id_alumno=alumno,
                                                                            nota=None)
                print("Evaluacion Creada")
            #evaluaciones = Evaluacion.
        else:
            print(form.errors)
        
        return redirect('sgn_app:ListarAlumno',id_gradoseccion=self.kwargs["id_gradoseccion"])   



# ----------------------------------------------
# Gestión de Docentes:
# ----------------------------------------------

# HU-28: Asignación de Grado/Sección con Materia
# Asignación de la materia que impartira el docente en un
# determinado grado y sección
class AsignacionClases(View):
    model = Gradoseccionmateria
    template_name = 'docente/asignacion_clases.html'

    def get_queryset(self):
        consultas = {}
        consultas = asignacionClases(self.request)
        consultas['Grado_seccion_materia'] = self.model.objects.all()
        consultas['Grado_seccion'] = Gradoseccion.objects.all()
        consultas['docentes'] = Docente.objects.all()
        return consultas

    def get_context_data(self, **kwargs):
        context = {}
        context = self.get_queryset()
        return context

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, self.get_context_data())

    def post(self, request, *args, **kwargs):
        if request.method == 'POST':
            opcion = request.POST['operacion']
            # Al ser la operacion un 1
            # Se crea un registro
            if opcion == '1':
                try:
                    id_grado_seccion = request.POST.get('idGradoSeccion', None)
                    id_docente = request.POST.get('idDocente', None)
                    nombre_materia = request.POST.get('nombreMateria', None)
                    existe_materia = Materia.objects.filter(
                        nombre_materia=nombre_materia, id_docente_id=id_docente).exists()
                    if existe_materia:
                        materias = Materia.objects.filter(
                            nombre_materia=nombre_materia, id_docente_id=id_docente)
                        existencia_grado_seccion_materia = Gradoseccionmateria.objects.filter(
                            id_materia__in=materias, id_gradoseccion_id=id_grado_seccion).exists()
                        if existencia_grado_seccion_materia:
                            messages.error(
                                request, 'No fue posible registrarlo porque se encontró un coincidencia en la base de datos')
                        else:
                            materia = Materia.objects.create(
                                id_docente_id=id_docente,
                                nombre_materia=nombre_materia
                            )

                            Gradoseccionmateria.objects.create(
                                id_gradoseccion_id=id_grado_seccion,
                                id_materia=materia
                            )
                    else:
                        materia = Materia.objects.create(
                            id_docente_id=id_docente,
                            nombre_materia=nombre_materia
                        )

                        Gradoseccionmateria.objects.create(
                            id_gradoseccion_id=id_grado_seccion,
                            id_materia=materia
                        )
                except Exception:
                    messages.error(
                        request, 'Ocurrio un error, introduzca datos válidos')
            else:
                # Al no ser lo anterior
                # Se edita un registro
                m = 0

        # ID de evaluación que deseas pasar a la URL
        return redirect('sgn_app:asignar_clases')


# Elimina la asignación seleccionada, que anteriormente
# se puedo crear en la clase AsignacionClases
def EliminarAsigacionClases(request, id):
    try:
        grado_seccion_materia = Gradoseccionmateria.objects.get(id_gradoseccionmateria = id)
        materia = Materia.objects.get(id_materia = grado_seccion_materia.id_materia.id_materia)
        grado_seccion_materia.delete()
        materia.delete()

    except Exception:
        messages.error(request, 'No es posible eliminar este registro')

    return redirect('sgn_app:asignar_clases')


# HU-29 Agrega Docentes
class CrearDocentes(View):
    model = Docente
    form_teacher = DocenteForm
    form_user = RegisterUserForm
    template_name = 'docente/crear_docente.html'

    def get_queryset(self):
        return self.model.objects.all()

    def get_context_data(self, **kwargs):
        context = {}
        context = asignacionClases(self.request)
        context['docente'] = self.get_queryset()
        context['docente_form'] = self.form_teacher
        context['user_form'] = self.form_user
        return context

    def get(self, request, *args, **kwargs):
        return render(request, self.template_name, self.get_context_data())

    def post(self, request, *args, **kwargs):
        form_teacher = self.form_teacher(request.POST)
        form_user = self.form_user(request.POST)
        if form_teacher.is_valid() and form_user.is_valid():
            form_teacher.save()
            form_user.save()
            messages.success(request, '!El docente se guardo exitosamente con su respectivas credenciales como usuario del sistema¡')
            return redirect('sgn_app:listado_docentes')
        else:
            messages.error(request, '¡Ocurrio un error!, verifica los datos ingresados anteriormente y vuelve a intentarlo')
            return render(request, self.template_name, self.get_context_data())


# HU-30: Editar Docentes
@login_required()
def EditarDocente(request, id):
    contexto = {}
    docente = get_object_or_404(Docente, numidentificacion=id)
    user = get_object_or_404(User, username=id)
    contexto = asignacionClases(request)
    
    if request.method == 'POST':
        form_teacher = DocenteForm(request.POST, instance=docente)
        form_user = UserChangeForm(request.POST, instance=user)
      
        if form_teacher.is_valid() and form_user.is_valid():
            form_teacher.save()
            form_user.save()
            messages.success(request, '!El docente se actualizo exitosamente¡')
            #update_session_auth_hash(request, user)  # Actualiza la sesión de autenticación
            return redirect('sgn_app:listado_docentes')
    else:
        form_teacher = DocenteForm(instance=docente)
        form_user = UserChangeForm(instance=user)
       
        contexto['docente_form'] = form_teacher
        contexto['user_form'] = form_user
        contexto['user_date_join'] = user.date_joined

    return render(request, 'docente/editar_docente.html', contexto)

@login_required()
def EditarDocenteContra(request, id):
    contexto = {}
    user = get_object_or_404(User, username=id)
    contexto = asignacionClases(request)
    
    if request.method == 'POST':
        password_form = SetPasswordForm(user, request.POST)  # Agrega el formulario de cambio de contraseña
       
        if password_form.is_valid():
            password_form.save()  # Guarda los cambios de contraseña
            #update_session_auth_hash(request, user)  # Actualiza la sesión de autenticación
            return redirect('sgn_app:listado_docentes')
    else:
        password_form = SetPasswordForm(user)  # Crea el formulario de cambio de contraseña
        contexto['password_form'] = password_form  # Agrega el formulario de cambio de contraseña al contexto
        contexto['user_date_join'] = user.date_joined

    return render(request, 'docente/editar_docente_contra.html',contexto)



# HU-31: Habilitar/Deshabilitar Docentes
# Vista para deshabilitar usuario
@login_required()
def deshabilitar_usuario(request, id):
    user = get_object_or_404(User, username=id)
    # Deshabilitar el usuario
    user.is_active = False
    user.save()
    return redirect('sgn_app:listado_docentes')


# Vista para habilitar usuario
@login_required()
def habilitar_usuario(request, id):
    user = get_object_or_404(User, username=id)
    # Deshabilitar el usuario
    user.is_active = True
    user.save()
    return redirect('sgn_app:listado_docentes')


# HU-32 Listar Docentes
class ListarDocentes(ListView):
    model = Docente
    template_name = 'docente/listar_docentes.html'
    context_object_name = 'docentes'
    queryset = model.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['users'] = User.objects.all()
        context.update(asignacionClases(self.request))
        return context

# ------------------------------------------


# HU-33: Crear Trimestre
def CrearTrimestre(request):
    form_trimestre = TrimestreForm(request.POST or None)
    contexto = {}
    contexto = asignacionClases(request)
    if request.method == 'POST':
        if form_trimestre.is_valid():
            form_trimestre.save()
            return redirect('sgn_app:crear_trimestre')
        else:
            messages.error(request, "¡El trimestre ya existe para ese año!")
    contexto['form_trimestre'] = form_trimestre
    return render(request, 'trimestre/crear_trimestre.html', contexto)


class Correcto(TemplateView):
    template_name = "trimestre/correcto.html"


# HU-34: Consultar Trimestres
class ListarTrimestres(ListView):
    model = Trimestre
    template_name = 'trimestre/listar_trimestres.html'
    context_object_name = 'trimestres'
    queryset = model.objects.all()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(asignacionClases(self.request))
        return context


# HU-35: Actualizar Trimestre
class ActualizarTrimestre(UpdateView):
    model = Trimestre
    template_name = "trimestre/actualizarTrim.html"
    form_class = TrimestreActualizarForm
    success_url = reverse_lazy('sgn_app:correcto')

# HU-36: Eliminar Trimestre


def EliminarTrimestre(request, id):
    try:
        trimestre = Trimestre.objects.get(id_trimestre=id)
        trimestre.delete()
    except Exception:
        messages.error(request, "¡No se puede eliminar el trimestre!")
    return redirect('sgn_app:listar_trimestres')


# HU-37: Listar Evaluaciones
class ListarEvaluaciones(ListView):
    model = Evaluacion
    template_name = 'evaluacion/listar_evaluaciones.html'
    context_object_name = 'evaluaciones'
    queryset = model.objects.all()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(asignacionClases(self.request))
        return context

# HU-38: Agregar Evaluación
def CrearEvaluacionAlumno(request):
    submitted = False
    if request.method == "POST":
        # form = EvaluacionForm(request.POST)
        form = EvaluacionForm(request.POST)
        try:
            form_valid = form.is_valid()
        except ValueError:
            form_valid = False
        if form_valid:
            evaluacion = form.save(commit=False)  # Evita guardar inmediatamente en la base de datos
            evaluacion.estado = 1  # Asigna el valor 1 a la columna "estado"
              # Ahora guarda el objeto con el nuevo valor en la base de datos
            # evaluacion = form.save()  # contiene los datos de la evaluacion que se acaba de crear
            # grado = form.cleaned_data['id_gradoseccionmateria'].id_gradoseccion.id_gradoseccion
            if evaluacion.id_trimestre is None:
                return HttpResponseRedirect('/estudiante/crear-eva-est')
            if evaluacion.porcentaje is None or evaluacion.porcentaje < 0 or evaluacion.porcentaje > 100:
                 messages.error(request, "El porcentaje de la evaluacion debe estar entre 0 y 100.")
                 return HttpResponseRedirect('/estudiante/crear-eva-est')
            if evaluacion is not None :
                grado = evaluacion.id_gradoseccionmateria.id_gradoseccion.id_gradoseccion
                alumno = Alumno.objects.filter(id_gradoseccion=grado)
                for e in alumno:
                    evaluacion.save()
                    evaalumno = Evaluacionalumno.objects.create(id_evaluacion=evaluacion, id_alumno=e, nota=None)
                    evaalumno.save()
                return HttpResponseRedirect('/estudiante/crear-eva-est?submitted=True')
        else:
            return HttpResponseRedirect('/estudiante/crear-eva-est')
                   
    else:
        form = EvaluacionForm
        # USER SUBMITTER THE FORM
        if 'submitted' in request.GET:
            submitted = True
        else:
            submitted = False
    context = {
        'form': form,
        'submitted': submitted,
    }
    context.update(asignacionClases(request))
    return render(request, 'estudiante/crear-evaluacion.html', context)


# HU-40: Editar Evaluación
class EvaluacionEditar(UpdateView):
    model = Evaluacion
    template_name = "evaluacion/editarEvaluacion.html"
    form_class = EvaluacionEditarForm
    success_url = reverse_lazy('sgn_app:listar_evaluaciones')

def evaluacion_editar_docente(request,idEvaluacion):
    contexto = asignacionClases(request)
    evaluacion = Evaluacion.objects.get(id_evaluacion = idEvaluacion)
    idgrado = evaluacion.id_gradoseccionmateria.id_gradoseccionmateria
    
    if request.method == 'POST':
        id_evaluacion = request.POST['id_evaluacion']
        nombre_evaluacion = request.POST['nombre_evaluacion']
        evaluacion.nombre_evaluacion = nombre_evaluacion
        evaluacion.save()
        return redirect('sgn_app:listar_evas_grado', idgrado=idgrado)
    else:
        contexto['evaluacion'] = evaluacion
        return render(request, 'evaluacion/editar_evaluacion_docente.html',contexto)

#HU-27 cargar alumnos en excel
@login_required
def excelAlumnos(request, id):
      # Semejante al home: para el navbar
    docente = Docente.objects.get(numidentificacion=request.user.username)
    materia = Materia.objects.filter(id_docente=docente)
    gradoseccionmateria = Gradoseccionmateria.objects.filter(id_materia__in=materia)
    gradoseccion = Gradoseccion.objects.filter(gradoseccionmateria__id_materia__id_docente=docente).distinct()
        
    if request.method == 'POST':
        try:
            if request.FILES.get('archivo_excel'):
                archivo_excel = request.FILES['archivo_excel']
            # Cargar el archivo Excel
                wb = load_workbook(archivo_excel)
                sheet = wb.active

            # Procesar cada fila del archivo Excel y guardar los datos en la sesión
                datos_archivo = []
                error=""  # Variable para rastrear errores en el Excel

                for row in sheet.iter_rows(min_row=2, values_only=True):
                    try:
                        nie, apellidos_alumno, nombres_alumno, sexo = row
                    except ValueError:
                        error = "Archivo de excel con formato incorrecto"
                        break  

                # Validar si algún campo está vacío
                    if not nie or not apellidos_alumno or not nombres_alumno or not sexo:
                        error="El excel no debe tener campos vacíos"
                        break  # Salir del bucle si hay un campo vacío en alguna fila
                
                # Convertir el valor de sexo a mayúsculas
                    sexo = sexo.upper()  # Esto convierte a mayúsculas independientemente de si estaba en minúsculas o mayúsculas

                # Validar que el valor de sexo sea "M" o "F"
                    if sexo not in ["M", "F"]:
                        error="Error, el sexo debe ser M o F."
                        break  # Salir del bucle si se encuentra un valor no válido

                    datos_archivo.append({'nie': nie, 'apellidos': apellidos_alumno, 'nombres': nombres_alumno,'sexo':sexo})

                if error:
                    messages.error(request, error)
                else:
                    # Guardar los datos en la sesión
                    request.session['datos_archivo'] = datos_archivo
                    return render(request, 'estudiante/verAlumnosExcel.html', {'id_gradoseccion': id, 'grado_seccion': gradoseccion, 'datos_archivo': datos_archivo, 'grado_seccion_materia': gradoseccionmateria})
            else:
                messages.error(request, "No se ha subido ningún excel")
        except:
            messages.error(request, "Archivo de excel con formato incorrecto")
    
    return render(request, 'estudiante/verAlumnosExcel.html', {'id_gradoseccion': id, 'grado_seccion': gradoseccion, 'grado_seccion_materia': gradoseccionmateria})


#Registra/importa los alumnos a la base de datos desde el excel
@login_required
def confirmar_importacion(request,id):
    # Recuperar los datos del archivo desde la sesión
    datos_archivo = request.session.get('datos_archivo', [])

    if request.method == 'POST':
        gradoseccion=Gradoseccion.objects.filter(id_gradoseccion=id).first()
        # Guardar los datos en la base de datos
        existing_alumno = None
        for data in datos_archivo:
            nie = data['nie']

            try:
                # Verificar si el alumno ya existe en la base de datos por su NIE
                existing_alumno = Alumno.objects.get(nie=nie, id_gradoseccion=gradoseccion)
                messages.error(request, f"El alumno con NIE {nie} ya está registrado.")
            except ObjectDoesNotExist:

                Alumno.objects.create(nie=data['nie'], apellidos_alumno=data['apellidos'], nombres_alumno=data['nombres'], id_gradoseccion=gradoseccion,estado=1,sexo=data['sexo'])
        if existing_alumno is None:
            messages.success(request,"Alumnos registrados correctamente")
        # Limpiar los datos en la sesión después de importar
        request.session['datos_archivo'] = None
    return redirect('sgn_app:ListarAlumno', id_gradoseccion=id)



#HU: 39 Habilitar/Deshabilitar Evaluaciones 
# Vista para deshabilitar Evaluación
@login_required
def deshabilitar_evaluacion(request, idgsm,idtri, id):
    evaluacion = get_object_or_404(Evaluacion, id_evaluacion=id)
    #Deshabilitar Evaluación
    evaluacion.estado = 0
    evaluacion.save()
    messages.success(request, f"La evaluación {evaluacion.nombre_evaluacion} ha sido deshabilitada con exito")
    # Redirigir al usuario a la página deseada
    return redirect(f'/evaluacion/listar-evas-grado/{idgsm}/?id_trimestre={idtri}') 


@login_required     
def habilitar_evaluacion(request,idgsm, idtri, id):
    evaluacion = get_object_or_404(Evaluacion, id_evaluacion=id)
    #Habilitar Evaluación
    evaluacion.estado = 1
    evaluacion.save()
    # Redirigir al usuario a la página deseada
    messages.success(request, f"La evaluación {evaluacion.nombre_evaluacion} ha sido habilitada con exito")
    return redirect(f'/evaluacion/listar-evas-grado/{idgsm}/?id_trimestre={idtri}')

#HU-17 Cuadro de honor
@login_required
def Elegir_trimestre(request):
    trimestres=Trimestre.objects.all()
    contexto={"trimestres":trimestres}
    contexto.update(asignacionClases(request))
    return render(request,"estudiante/elegir-trimestre.html",contexto)

@login_required
def cuadro_honor(request):
    trimestre_seleccionado = request.POST.get('trimestreSelect')
    alumnos=Alumno.objects.filter(estado=1)
    trimestre=Trimestre.objects.filter(id_trimestre=trimestre_seleccionado).first()
    cuadroHonor=[]
    for alumno in alumnos:
        promedioAlumno=0.00
        materiasAlumno=Gradoseccionmateria.objects.filter(id_gradoseccion=alumno.id_gradoseccion)     
        for materia in materiasAlumno:
            promedioMateria=0.00

            evaluacionesAlumno=Evaluacionalumno.objects.filter(
                id_alumno=alumno.id_alumno,
                id_evaluacion__id_gradoseccionmateria__id_materia=materia.id_materia,
                id_evaluacion__id_trimestre=trimestre_seleccionado,
                id_evaluacion__estado=1)
            for evaluacion in evaluacionesAlumno:
                if evaluacion.nota:
                    promedioMateria+=evaluacion.nota*(evaluacion.id_evaluacion.porcentaje/100)
            promedioAlumno+=promedioMateria/len(materiasAlumno)
            
        if promedioAlumno==9.99:
            promedioAlumno=10
        if promedioAlumno>=8:
            cuadroHonor.append({'alumno': alumno, 'promedioAlumno':promedioAlumno})
            
    contexto={"cuadroHonor":cuadroHonor,"trimestre":trimestre}
    contexto.update(asignacionClases(request))
    return render(request, "estudiante/cuadro_honor.html",contexto)
#HU-18
def GestionTelas(request):
    context = {}

    yardasPorNivel = {
        "parvularia":
        {
            "nivel": Gradoseccion.Nivel.PARVULARIA.label,
            "blusa": Decimal(0.75),
            "falda": round(Decimal(0.60),2),
            "camisa": Decimal(0.75),
            "pantaloncorto": Decimal(0.75),
            "pantalon": Decimal(0.75)
        },
        "primerciclo":
        {
            "nivel": Gradoseccion.Nivel.PRIMERCICLO.label,
            "blusa": Decimal(1.00),
            "falda": Decimal(0.75),
            "camisa": Decimal(1.00),
            "pantalon": Decimal(1.00)
        },
        "segundociclo":
        {
            "nivel": Gradoseccion.Nivel.SEGUNDOCICLO.label,
            "blusa": Decimal(1.25),
            "falda": Decimal(1.00),
            "camisa": Decimal(1.25),
            "pantalon": Decimal(1.25)
        },
        "tercerciclo":
        {
            "nivel": Gradoseccion.Nivel.TERCERCICLO.label,
            "blusa": Decimal(1.50),
            "falda": Decimal(1.25),
            "camisa": Decimal(1.50),
            "pantalon": Decimal(1.50)
        },
    }

    

    #TABLA TOTAL ALUMNOS CENTRO ESCOLAR
    resumen_alumnos = [
        [
            "Parvularia",
            Alumno.objects.filter(id_gradoseccion__nivel=Gradoseccion.Nivel.PARVULARIA,sexo=Alumno.Sexo.MASCULINO).count(),
            Alumno.objects.filter(id_gradoseccion__nivel=Gradoseccion.Nivel.PARVULARIA,sexo=Alumno.Sexo.FEMENINO).count(),
        ],
        [
            "Básica",
            Alumno.objects.filter(sexo=Alumno.Sexo.MASCULINO).count() - Alumno.objects.filter(id_gradoseccion__nivel=Gradoseccion.Nivel.PARVULARIA,sexo=Alumno.Sexo.MASCULINO).count(),
            Alumno.objects.filter(sexo=Alumno.Sexo.FEMENINO).count() - Alumno.objects.filter(id_gradoseccion__nivel=Gradoseccion.Nivel.PARVULARIA,sexo=Alumno.Sexo.FEMENINO).count(),
             
        ]
    ]
        
    #TABLA PRENDAS PARVULARIA
    camisa_p = yardas(Alumno.Sexo.MASCULINO,Gradoseccion.Nivel.PARVULARIA,yardasPorNivel,"parvularia","camisa")
    blusa_p = yardas(Alumno.Sexo.FEMENINO,Gradoseccion.Nivel.PARVULARIA,yardasPorNivel,"parvularia","blusa")
    short_p = yardas(Alumno.Sexo.MASCULINO,Gradoseccion.Nivel.PARVULARIA,yardasPorNivel,"parvularia","pantaloncorto")
    falda_p = yardas(Alumno.Sexo.FEMENINO,Gradoseccion.Nivel.PARVULARIA,yardasPorNivel,"parvularia","falda")
    
    #TABLA PRENDAS PRIMERCICLO
    camisa_pc = yardas(Alumno.Sexo.MASCULINO,Gradoseccion.Nivel.PRIMERCICLO,yardasPorNivel,"primerciclo","camisa")
    blusa_pc = yardas(Alumno.Sexo.FEMENINO,Gradoseccion.Nivel.PRIMERCICLO,yardasPorNivel,"primerciclo","blusa")
    falda_pc = yardas(Alumno.Sexo.FEMENINO,Gradoseccion.Nivel.PRIMERCICLO,yardasPorNivel,"primerciclo","falda")
    pantalon_pc = yardas(Alumno.Sexo.MASCULINO,Gradoseccion.Nivel.PRIMERCICLO,yardasPorNivel,"primerciclo","pantalon")
    print("helfjalsdf ",falda_pc)
    #TABLA PRENDAS SEGUNDOCICLO
    camisa_sc = yardas(Alumno.Sexo.MASCULINO,Gradoseccion.Nivel.SEGUNDOCICLO,yardasPorNivel,"segundociclo","camisa")
    blusa_sc = yardas(Alumno.Sexo.FEMENINO,Gradoseccion.Nivel.SEGUNDOCICLO,yardasPorNivel,"segundociclo","blusa")
    falda_sc = yardas(Alumno.Sexo.FEMENINO,Gradoseccion.Nivel.SEGUNDOCICLO,yardasPorNivel,"segundociclo","falda")
    pantalon_sc = yardas(Alumno.Sexo.MASCULINO,Gradoseccion.Nivel.SEGUNDOCICLO,yardasPorNivel,"segundociclo","pantalon")

    #TABLA PRENDAS TERCERCICLO
    camisa_tc = yardas(Alumno.Sexo.MASCULINO,Gradoseccion.Nivel.TERCERCICLO,yardasPorNivel,"tercerciclo","camisa")
    blusa_tc = yardas(Alumno.Sexo.FEMENINO,Gradoseccion.Nivel.TERCERCICLO,yardasPorNivel,"tercerciclo","blusa")
    falda_tc = yardas(Alumno.Sexo.FEMENINO,Gradoseccion.Nivel.TERCERCICLO,yardasPorNivel,"tercerciclo","falda")
    pantalon_tc = yardas(Alumno.Sexo.MASCULINO,Gradoseccion.Nivel.TERCERCICLO,yardasPorNivel,"tercerciclo","pantalon")

    parvularia = [camisa_p,blusa_p,short_p,falda_p]
    primer_ciclo = [camisa_pc,blusa_pc,falda_pc,pantalon_pc]
    segundo_ciclo = [camisa_sc,blusa_sc,falda_sc,pantalon_sc]
    tercer_ciclo = [camisa_tc,blusa_tc,falda_tc,pantalon_tc]

    #TABLA TOTAL POR COLOT DE TELA
    celeste = camisa_p[2] + blusa_p[2]
    blanca = camisa_pc[2] + blusa_pc[2] + camisa_sc[2] + blusa_sc[2] + camisa_tc[2] + blusa_tc[2]
    azul = falda_tc[2] + pantalon_tc[2] + falda_sc[2] + pantalon_sc[2] +falda_pc[2] + pantalon_pc[2] + short_p[2] + falda_p[2]
    beige = 0

    yardas_colores = [
        ["Celeste",celeste],
        ["Blanca",blanca],
        ["Azul",azul]
    ]
    
    context["parvularia"] = parvularia
    context["primer_ciclo"] = primer_ciclo
    context["segundo_ciclo"] = segundo_ciclo
    context["tercer_ciclo"] = tercer_ciclo
    context["resumen_alumnos"] = resumen_alumnos
    context["yardas_por_nivel"] = yardasPorNivel
    context["yardas_colores"] = yardas_colores

    return render (request,'gestion-tela/telas.html',context)


def yardas(sexo_,nivel_,yardas_,nivel_2,nombre_prenda):
    txt = ""
    
    query = Alumno.objects.filter(id_gradoseccion__nivel=nivel_,sexo=sexo_).count()
    
    total_yardas = yardas_[nivel_2][nombre_prenda] * query

    if nombre_prenda=="pantaloncorto":
        txt = "Pantalón Corto"
    elif nombre_prenda == "pantalon":
        txt = "Pantalón"
    else: 
        txt =nombre_prenda.title()
    
    lista = [txt,query,round(Decimal(total_yardas),2)]
    
    return lista